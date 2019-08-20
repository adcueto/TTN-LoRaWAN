
import openpyxl

dest_filename = "/Users/turulo/Developer/Python/Hello/LoRaWAN.xlsx"
rowc=2; columnc=0

doc = openpyxl.load_workbook(filename=dest_filename)
general = doc.worksheets[0]
general['A2'] = "hola"
general.cell(row=rowc, column=2).value = 222
rowc=rowc+1
general.cell(row=rowc, column=2).value = 223
rowc=rowc+1
general.cell(row=rowc, column=2).value = 224
print(general.cell(row=2, column=2).value)
doc.save(filename = dest_filename)
